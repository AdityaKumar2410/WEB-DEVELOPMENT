from flask import Flask, request, render_template
from openpyxl import load_workbook, Workbook

app = Flask(__name__)

# Ensure the Excel file exists or create it
def ensure_excel_file(filename):
    try:
        load_workbook(filename)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Username", "Email", "Phone", "Gender", "Password"])  # Add headers
        wb.save(filename)

@app.route("/")
def index():
    return render_template("index.html")  # Your updated HTML form

@app.route("/submit", methods=["POST"])
def submit():
    username = request.form["username"]
    email = request.form["email"]
    phone = request.form["phone"]
    gender = request.form["gender"]
    password = request.form["password"]

    filename = "user_data.xlsx"
    ensure_excel_file(filename)
    wb = load_workbook(filename)
    ws = wb.active
    ws.append([username, email, phone, gender, password])  # Add new data
    wb.save(filename)
    return "Data has been saved successfully!"

if __name__ == "__main__":
    app.run(debug=True)
